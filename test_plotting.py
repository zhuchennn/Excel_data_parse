
''' 
@author: Zhu Chen
'''

import openpyxl, os
import pandas as pd
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime
import tkinter as tk

#input: the overarching file path of where all the relevant excel files are stored
#output: a dictionary with key: file path of the specific xlsx file and value: list of filenames of xlsx files
def loop_directory(direct: str):
    dic = {}
    for (root, dirs, files) in os.walk(direct, topdown=True):
        lst_files = []
        for i in files:
            if i.endswith('.xlsx') and not i.startswith('~$'):
                lst_files.append(i)
        dic[root] = lst_files
    return dic

def get_root_file(root, filename):
    return os.path.join(str(root), filename) #<---- This joins the directory(str) and filename(str) and returns a str

#input: a dictionary the keyword being the filename
#output: the root of the file and the filename in a 2 index list
def get_file_path(dic: dict, keyword: str):  #need to get a specific file with a specific input
    for root, filenames in dic.items():
        for filename in filenames:
            if keyword in filename:
                file_path = Path(get_root_file(root, filename))
                f = open(file_path)
                return [root, filename]

def get_key(dic: dict, val: str):
    for key, value in dic.items():
         if val == value:
             return key

def get_para(cpk_sheet, dic_main: dict, para: str, config: str):
    for row in cpk_sheet.iter_rows(min_col = 3, max_col = 3, min_row = 2,max_row = maxrow):
        length = 0
        if para in row[0].value:
            if para not in dic_main['Parameter'].values():
                if dic_main['Parameter'] != {}:
                    length = max(dic_main['Parameter'].keys()) +1
                else:
                    length = 1
                dic_main['Parameter'][length] = para
            else:
                length = get_key(dic_main['Parameter'], para)
            cpk = round(cpk_sheet['M' + str(row[0].row)].value, 2)
            dic_main[config][length] = cpk
            
            
            
window1 = tk.Tk()
canvas1 = tk.Canvas(window1, width=400, height=355, relief='raised')
canvas1.pack()

title = tk.Label(window1, text='Input Window')
title.config(font=('helvetica', 14))
canvas1.create_window(200, 25, window=title)

title1 = tk.Label(window1, text = 'File Path Input')
title1.config(font=('helvetica', 10))
canvas1.create_window(200, 80, window=title1)
direct = tk.Entry(window1)
canvas1.create_window(200, 115, window=direct)

title2 = tk.Label(window1, text= 'File Name Input')
title2.config(font=('helvetica', 10))
canvas1.create_window(200, 155, window=title2)
filename = tk.Entry(window1)
canvas1.create_window(200, 190, window=filename)

title3 = tk.Label(window1, text= 'Graph Header Input')
title3.config(font=('helvetica', 10))
canvas1.create_window(200, 230, window=title3)
plot_header = tk.Entry(window1)
canvas1.create_window(200, 265, window=plot_header)

def close_window():
    global direct
    direct = direct.get()
    global filename
    filename = filename.get()
    global plot_header
    plot_header = plot_header.get()
    window1.destroy()

button1 = tk.Button(text= 'confirm', command = close_window)
canvas1.create_window(200, 305, window=button1)

window1.mainloop()

#direct = r'C:\Users\zzh2\Desktop\Zhu_Chen\Python_script\sherlock reports'
#filename = r'Book3.xlsx' #################change_____________________________________
#dic_input = {}
dic_files = loop_directory(direct)
#print(dic_files)
os.chdir(direct)
root = get_file_path(dic_files, filename)[0]
file = get_file_path(dic_files, filename)[1] #actually redundant
os.chdir(root)
data = pd.ExcelFile(file) #helps make the data from excel file readable by python
ps = openpyxl.load_workbook(file) #similar to above, makes excel file readable to python

sheet = ps[data.sheet_names[0]]
maxrow = sheet.max_row
maxcol = sheet.max_column
x = ps.active
dic_input = {}
lst_config = [] #e.g. ['ad','bd','cd','ed']
lst_para = [] #e.g. ['DKCT::DARK_AsyncM2_1st_CH1', 'DKCT::DARK_AsyncM2_1st_CH5', 'DKCT::DARK_AsyncM2_Avg_CH1']
for row in x.iter_rows(max_col=2, max_row=maxrow):
    #dic_input[row[0].value] = row[1].value
    # row is in a tuple format e.g. ('ad', 'DKCT::DARK_AsyncM2_1st_CH1')
    if isinstance(row[0].value, str) or isinstance(row[0].value, int):
        lst_config.append(row[0].value) #config
    if isinstance(row[1].value, str):
        lst_para.append(row[1].value) #para


dic_main ={}
dic_main.setdefault('Parameter', {})
glob_header = ['Parameter']
for config in lst_config:
    glob_header.append(config)
    for root, filenames in dic_files.items():
        for filename in filenames:
            print(filename)

    root = get_file_path(dic_files, config)[0]
    file = get_file_path(dic_files, config)[1]
    os.chdir(root)
    data = pd.ExcelFile(file)
    dic_main.setdefault(config, {})
    print(data.sheet_names)
    for i in data.sheet_names:
        if 'Overall_Cpk' in i:
            s= i

    df = data.parse(s)
    ps = openpyxl.load_workbook(file)
    sheet = ps[s]
    maxrow = sheet.max_row
    maxcol = sheet.max_column
    ps.active = ps[s]
    cpk_sheet = ps.active 
    for para in lst_para:
        get_para(cpk_sheet, dic_main, para, config)

base_cpk_indicator = max(dic_main['Parameter'].keys())+1

for k, v in dic_main.items():
    if k != 'Parameter':
        dic_main[k][base_cpk_indicator] = 1.33
    else:
        dic_main[k][base_cpk_indicator] = 'Base CPK'

df = pd.DataFrame(dic_main, columns = glob_header)
df.set_index(['Parameter'], inplace=True)

unique=datetime.datetime.now().strftime('%m%d%H%M')
new_name = direct + r'\Test_CPK_' + str(unique) + '_' + plot_header + r'.xlsx'
writer = pd.ExcelWriter(new_name, engine='xlsxwriter')

sheetname = 'Sheet1'
df.to_excel(writer, sheet_name = sheetname)
workbook = writer.book
worksheet = writer.sheets[sheetname]

#plot
col = 1 #start col
max_col = len(dic_main.keys())-1 
(max_row, maximum_col) = df.shape
chart =  workbook.add_chart({'type': 'line'})
for r in range(1, max_row+1):
    chart.add_series({
            'name':       [sheetname, r, 0],
            #format of [start_row, start_col, end_row, end col]
            'categories': [sheetname, 0, col, 0, max_col], #obtains labels aka x axis 
            'values':     [sheetname, r, col, r, max_col], #obtains data, aka y axis
    })

chart.set_title({'name': plot_header}) #sets title of the charts
chart.set_x_axis({'name': 'Configs'}) #sets name of the x axis
chart.set_y_axis({'name': 'CPK', 'major_gridlines': {'visible': False}}) #sets name of the y axis
chart.set_table({'show_keys': True}) 

letter_chart = get_column_letter(maximum_col +6)
ins = letter_chart + str(3) 
worksheet.insert_chart(ins, chart)

writer.save()
print('Done')


