from __future__ import absolute_import
from cmath import nan
import openpyxl, os
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd 
import tkinter as tk
import datetime

##########################################################################################################################

# Copyright (c) 2010-2021 openpyxl

from io import BytesIO
from warnings import warn

from openpyxl.xml.functions import fromstring
from openpyxl.xml.constants import IMAGE_NS
from openpyxl.packaging.relationship import get_rel, get_rels_path, get_dependents
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.drawing.image import Image, PILImage
from openpyxl.chart.chartspace import ChartSpace
from openpyxl.chart.reader import read_chart

def find_images(archive, path):
    """
    Given the path to a drawing file extract charts and images

    Ingore errors due to unsupported parts of DrawingML
    """

    src = archive.read(path)
    tree = fromstring(src)
    try:
        drawing = SpreadsheetDrawing.from_tree(tree)
    except TypeError:
        warn("DrawingML support is incomplete and limited to charts and images only. Shapes and drawings will be lost.")
        return [], []

    rels_path = get_rels_path(path)
    deps = []
    if rels_path in archive.namelist():
        deps = get_dependents(archive, rels_path)

    charts = []
    for rel in drawing._chart_rels:
        cs = get_rel(archive, deps, rel.id, ChartSpace)
        chart = read_chart(cs)
        chart.anchor = rel.anchor
        charts.append(chart)

    images = []
    if not PILImage: # Pillow not installed, drop images
        return charts, images

    for rel in drawing._blip_rels:
        dep = deps[rel.embed]
        if dep.Type == IMAGE_NS:
            try:
                image = Image(BytesIO(archive.read(dep.target)))
            except OSError:
                msg = "The image {0} will be removed because it cannot be read".format(dep.target)
                warn(msg)
                continue
            except KeyError:
                warn('Missing image')
                continue
            if image.format.upper() == "WMF": # cannot save
                msg = "{0} image format is not supported so the image is being dropped".format(image.format)
                warn(msg)
                continue
            image.anchor = rel.anchor
            images.append(image)
    return charts, images
####################################################################################################################################





#input: the overarching file path of where all the relevant excel files are stored
#output: a dictionary with key: file path of the specific xlsx file and value: filename of xlsx file
def loop_directory(direct: str):
    dic = {}
    for (root, dirs, files) in os.walk(direct, topdown=True):
        lst_files = []
        for i in files:
            if i.endswith('.xlsx') and not i.startswith('~$'):
                lst_files.append(i)
        dic[root] = lst_files
    return dic

def get_root_file(root, filename):
    return os.path.join(str(root), filename) #<---- This joins the directory(str) and filename(str) and returns a str

#input: a dictionary the keyword being the filename
#output: the root of the file and the filename in a 2 index list
def get_file_path(dic: dict, keyword: str):  #need to get a specific file with a specific input
    for root, filenames in dic.items():
        for filename in filenames:
            if keyword in filename:
                file_path = Path(get_root_file(root, filename))
                f = open(file_path)
                return [root, filename]

#input: the keys of the cell you want the column for and the cell you want the row for
#*i is automatically in the form of a tuple and is used to modify the row coordinates
#output: the coordinate of the new required cell
def get_new_cell(col_cell, row_cell, *i): #input is not a string
    if i:
        return get_column_letter(col_cell.column) + str(row_cell.row + i[0]) #output is a string
    return get_column_letter(col_cell.column) + str(row_cell.row)

#finds the coordinates of the cell with the heading 3. Measurement Data
def get_measurement_data(sheet): #contains fixed input
    for rows in sheet:
        for col in rows:
            if col.value == '3. Measurement Data': #MUST BE THIS FIXED INPUT
                return col #does not return a string

#compile column row into a dictionary with cells as keys and cell content as value
#mbo_pbo is the name of the active sheet
def get_header_cell(mbo_pbo, build: str): 
    dic = {}
    for col in mbo_pbo.iter_cols(min_col=1, max_col= maxcol , min_row= start.row + 2, max_row=start.row + 2):
        dic[col[0]] = col[0].value
    #print(dic)
    for key, value in dic.items():
        if type(value) == str and build in value: #MODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFY
            return key #not a string

def get_process(mbo_pbo):
    for row in mbo_pbo.iter_rows(min_col= 2, max_col=2, max_row=maxrow):
        if row[0].value != None and row[0].value != ' ' and row[0].value == 'Process':
            y = row[0].row+1
            x = get_column_letter(row[0].column)
            return mbo_pbo[x + str(y)].value

def get_key(dic: dict, val: str):
    for key, value in dic.items():
         if val == value:
             return key

#obtians main dic with ALL parameters
#also obtains cpk dic with ONLY cpk values, spc items and process names
def get_overall_dic(mbo_pbo, build: str, dic_global: dict):
    dic_local = {} #contains the local dict of the file in the format of {process: value}

    
    #obtains the 6 values in a list and puts them into a dict with key=SPC Item and value= list of the 6 values
    for row in mbo_pbo.iter_rows(min_col= 3, max_col=3, min_row= start.row + 3, max_row=maxrow):
        
        if row[0].value != None and row[0].value != ' ' and row[0].value == 'CPK': #remove 2nd boolean if you want other values too
            #1st boolean ensures no None type cells
            #2nd boolean ensrues no empty cells
            #3rd boolean finds the cell containing the CPK value
            lst_item = []

            cpk_name_cell = row[0] #for the cpk row, aka number, this is a whole coordinate, not just row number

            cell = 'B' + str(cpk_name_cell.row) #cell for SPC Item
            cell2 = get_header_cell(mbo_pbo, build) #for the column, aka letter
            N_cell = get_new_cell(cell2, cpk_name_cell, -5)
            max_cell = get_new_cell(cell2, cpk_name_cell, -4)
            min_cell = get_new_cell(cell2, cpk_name_cell, -3)
            mean_cell = get_new_cell(cell2, cpk_name_cell, -2)
            std_dev_cell = get_new_cell(cell2, cpk_name_cell, -1)
            cpk_cell = get_new_cell(cell2, cpk_name_cell)
            LSL_cell = get_new_cell(cell2, cpk_name_cell, 4)
            USL_cell = get_new_cell(cell2, cpk_name_cell, 5)
            target_cell = get_new_cell(cell2, cpk_name_cell, 6)

            lst_item.append(mbo_pbo[N_cell].value) #0
            lst_item.append(mbo_pbo[min_cell].value) #1
            lst_item.append(mbo_pbo[mean_cell].value) #2
            lst_item.append(mbo_pbo[max_cell].value) #3
            lst_item.append(mbo_pbo[std_dev_cell].value) #4
            lst_item.append(mbo_pbo[cpk_cell].value) #5
            lst_item.append(mbo_pbo[LSL_cell].value) #6 
            lst_item.append(mbo_pbo[USL_cell].value)#7
            lst_item.append(mbo_pbo[target_cell].value) #8

            dic_local[mbo_pbo[cell].value] = lst_item
    if build+ ' 5' not in cpk_head:
        cpk_head.append(build+ ' 5')
    cpk_dic.setdefault(build + ' 5', {0: header2_col[5]})

    lst = []
    for i in range(len(header2_col)):
        #new = header[1] + ' ' + str(i)
        new = build + ' ' + str(i) #MODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFY
        lst.append(new)
        if new not in glob_header:
            glob_header.append(new)

        dic_global.setdefault(new, {0: header2_col[i]}) #for all values i.e. N, mean, min, max, cpk, std dev

    #set the 6 values into the respective spc rows
    for name, lst_item in dic_local.items():
        

        if name not in dic_global['SPC Item'].values(): #for new SPC items
            length = max(dic_global['SPC Item'].keys())
            length2 = max(cpk_dic['SPC Item'].keys()) +1

            length += 1
            dic_global['SPC Item'][length] = name
            
            if dic_global[lst[0]][0] == 'N':
                dic_global[lst[0]][length] = lst_item[0] #N

            if dic_global[lst[1]][0] == 'Min':
                dic_global[lst[1]][length] = lst_item[1] #min

            if dic_global[lst[2]][0] == 'Mean':
                dic_global[lst[2]][length] = lst_item[2] #mean

            if dic_global[lst[3]][0] == 'Max':
                dic_global[lst[3]][length] = lst_item[3] #max
                
            if dic_global[lst[4]][0] == 'Std Dev':
                dic_global[lst[4]][length] = lst_item[4] #Std dev
                
            if dic_global[lst[5]][0] == 'CPK': #need do for all 6
                if lst_item[5] != '-' and lst_item[5] != '' and lst_item[5] != None:
                    lst_item[5] = round(lst_item[5], 2)
                    cpk_dic['SPC Item'][length2] = name ##############HMMMMMMMMMMMMMMMMMMMMMM?###############
                    cpk_dic[lst[5]][length2] = lst_item[5] #CPK for the CPK ONLY data
                dic_global[lst[5]][length] = lst_item[5] #CPK
            
            dic_global['LSL'][length] = lst_item[6]
            dic_global['Target'][length] = lst_item[8]
            dic_global['USL'][length] = lst_item[7]

        else: #for different builds e.g. LS3 vs US3 etc. where SPC item name is alr there
            ind = get_key(dic_global['SPC Item'], name)#finding index of the SPC items
            ind2 = get_key(cpk_dic['SPC Item'], name)
            if dic_global[lst[0]][0] == 'N':
                dic_global[lst[0]][ind] = lst_item[0]

            if dic_global[lst[1]][0] == 'Min':
                dic_global[lst[1]][ind] = lst_item[1]

            if dic_global[lst[2]][0] == 'Mean':
                dic_global[lst[2]][ind] = lst_item[2]

            if dic_global[lst[3]][0] == 'Max':
                dic_global[lst[3]][ind] = lst_item[3]

            if dic_global[lst[4]][0] == 'Std Dev':
                dic_global[lst[4]][ind] = lst_item[4]

            if dic_global[lst[5]][0] == 'CPK': #need do for all 6
                if lst_item[5] != '-' and lst_item[5] != '' and lst_item[5] != None:
                    lst_item[5] = round(lst_item[5], 2)
                    cpk_dic[lst[5]][ind2] = lst_item[5]
                dic_global[lst[5]][ind] = lst_item[5]

            dic_global['LSL'][ind] = lst_item[6]
            dic_global['Target'][ind] = lst_item[8]
            dic_global['USL'][ind] = lst_item[7]

    return dic_global
    #return(x)

def get_overall_liner_attach_dic(mbo_pbo, build: str, dic_global: dict):
    dic_local = {}
    #obtains the 6 values in a list and puts them into a dict with key=SPC Item and value= list of the 6 values
    for row in mbo_pbo.iter_rows(min_col= 3, max_col=3, min_row= start.row + 3, max_row=maxrow):
        
        #x.append(row)
        if row[0].value != None and row[0].value != ' ' and row[0].value == 'Min': #remove 2nd boolean if you want other values too
            #1st boolean ensures no None type cells
            #2nd boolean ensrues no empty cells
            #3rd boolean finds the cell containing the CPK value
            lst_item = []
            cpk_name_cell = row[0] #for the cpk row, aka number, this is a whole coordinate, not just row number

            cella = 'B' + str(cpk_name_cell.row) #cell for SPC Item
            cellb = 'B' + str(cpk_name_cell.row + 1)
            spcitem = mbo_pbo[cella].value + mbo_pbo[cellb].value
            cell2 = get_header_cell(mbo_pbo, build) #for the column, aka letter
            N_cell = get_new_cell(cell2, cpk_name_cell, -2)
            max_cell = get_new_cell(cell2, cpk_name_cell, -1)
            min_cell = get_new_cell(cell2, cpk_name_cell)
            mean_cell = get_new_cell(cell2, cpk_name_cell, 1)
            LSL_cell = get_new_cell(cell2, cpk_name_cell, 2)
            USL_cell = get_new_cell(cell2, cpk_name_cell, 3)
            target_cell = get_new_cell(cell2, cpk_name_cell, 4)

            lst_item.append(mbo_pbo[N_cell].value) #0
            lst_item.append(mbo_pbo[min_cell].value) #1
            lst_item.append(mbo_pbo[mean_cell].value) #2
            lst_item.append(mbo_pbo[max_cell].value) #3
            lst_item.append(mbo_pbo[LSL_cell].value) #4 
            lst_item.append(mbo_pbo[USL_cell].value)#5
            lst_item.append(mbo_pbo[target_cell].value) #6

            dic_local[spcitem] = lst_item
    lst = []
    for i in range(len(header2_col)):
        #new = header[1] + ' ' + str(i)
        new = build + ' ' + str(i) #MODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFYMODIFY
        lst.append(new)
        if new not in glob_header:
            glob_header.append(new)

        dic_global.setdefault(new, {0: header2_col[i]})

    for name, lst_item in dic_local.items():
        if name not in dic_global['SPC Item'].values(): #for new SPC items
            length = max(dic_global['SPC Item'].keys())
            length2 = max(cpk_dic['SPC Item'].keys()) +1

            length += 1
            dic_global['SPC Item'][length] = name
            
            if dic_global[lst[0]][0] == 'N':
                dic_global[lst[0]][length] = lst_item[0] #N

            if dic_global[lst[1]][0] == 'Min':
                dic_global[lst[1]][length] = lst_item[1] #min

            if dic_global[lst[2]][0] == 'Mean':
                dic_global[lst[2]][length] = lst_item[2] #mean

            if dic_global[lst[3]][0] == 'Max':
                dic_global[lst[3]][length] = lst_item[3] #max
            
            dic_global['LSL'][length] = lst_item[4]
            dic_global['Target'][length] = lst_item[6]
            dic_global['USL'][length] = lst_item[5]
            
            

        else: #for different builds e.g. LS3 vs US3 etc. where SPC item name is alr there
            ind = get_key(dic_global['SPC Item'], name)#finding index of the SPC items

            if dic_global[lst[0]][0] == 'N':
                dic_global[lst[0]][ind] = lst_item[0]

            if dic_global[lst[1]][0] == 'Min':
                dic_global[lst[1]][ind] = lst_item[1]

            if dic_global[lst[2]][0] == 'Mean':
                dic_global[lst[2]][ind] = lst_item[2]

            if dic_global[lst[3]][0] == 'Max':
                dic_global[lst[3]][ind] = lst_item[3]

            dic_global['LSL'][ind] = lst_item[4]
            dic_global['Target'][ind] = lst_item[6]
            dic_global['USL'][ind] = lst_item[5]
    
    return dic_global

#OBTAIN INPUTS
window1 = tk.Tk()
canvas1 = tk.Canvas(window1, width=400, height=300, relief='raised')
canvas1.pack()

title = tk.Label(window1, text='Data Extraction Inputs')
title.config(font=('helvetica', 14))
canvas1.create_window(200, 25, window=title)

title1 = tk.Label(window1, text = 'File Path Input')
title1.config(font=('helvetica', 10))
canvas1.create_window(200, 80, window=title1)
direct = tk.Entry(window1)
canvas1.create_window(200, 115, window=direct)

title2 = tk.Label(window1, text= 'File Name Input')
title2.config(font=('helvetica', 10))
canvas1.create_window(200, 155, window=title2)
filename = tk.Entry(window1)
canvas1.create_window(200, 190, window=filename)

def close_window():
    global direct
    direct = direct.get()
    global filename
    filename = filename.get()
    window1.destroy()

button1 = tk.Button(text= 'confirm', command = close_window)
canvas1.create_window(200, 230, window=button1)

window1.mainloop()

os.chdir(direct)

    
#obtain filnames and file keywords from the excel file
dic_input = {}
dic_files = loop_directory(direct)

root = get_file_path(dic_files, filename)[0]
file = get_file_path(dic_files, filename)[1]
os.chdir(root)
data = pd.ExcelFile(file)
df = data.parse(data.sheet_names[0])
ps = openpyxl.load_workbook(file)
sheet = ps[data.sheet_names[0]]
maxrow = sheet.max_row
maxcol = sheet.max_column
x = ps.active
for row in x.iter_rows(max_col=2, max_row=maxrow):
    dic_input[row[0].value] = row[1].value

#MAIN LOOP to obtain the table
cpk_dic = {}
cpk_head = ['SPC Item', 'Process']
cpk_dic.setdefault('SPC Item', {0: nan})
cpk_dic.setdefault('Process', {0: nan})
process_length_lst = []
process_lst = []

dic_main ={}
index_header = [] #header for rows
glob_header = ['SPC Item', 'Process', 'LSL', 'Target', 'USL'] #header for columns

header1_col = ['SPC Item', 'FAI', 'SPC', 'LSL', 'Target', 'USL']
header2_col = ['N', 'Min', 'Mean', 'Max', 'Std Dev', 'CPK']
dic_main.setdefault('LSL', {0: nan})
dic_main.setdefault('Target', {0: nan})
dic_main.setdefault('USL', {0: nan})
for f_name, build in dic_input.items():

    header = [glob_header[0]]
    header.append(build)
    
    root = get_file_path(dic_files, f_name)[0]
    file = get_file_path(dic_files, f_name)[1]
    os.chdir(root)

    data = pd.ExcelFile(file)
    print(data.sheet_names)
    for i in data.sheet_names:
        if 'MBO PBO Summary' in i:
            s= i
    #df = data.parse(data.sheet_names[2])
    df = data.parse(s)

    print(file)
    print(df.head(10))
    ps = openpyxl.load_workbook(file)
    #sheet = ps[data.sheet_names[2]]
    sheet = ps[s]
    maxrow = sheet.max_row
    maxcol = sheet.max_column

    
    start = get_measurement_data(sheet)
    m_data = get_column_letter(start.column) + str(start.row + 2)

    

    print(start.coordinate)
    ps.active = ps[s]
    mbo_pbo = ps.active #input for get_header_cell()
    process = get_process(mbo_pbo)
    if process not in index_header:
        index_header.append(process)

    dic_main.setdefault('SPC Item', {0: nan})
    dic_main.setdefault('Process', {0: nan})
    #set Process names into dict
    if max(dic_main['Process'].keys()) == 0 and process not in dic_main['Process'].values():
        print(process)
        dic_main['Process'][1] = process #process is global variable
        cpk_dic['Process'][1] = process
        process_lst .append(process)
    elif process not in dic_main['Process'].values():
        print(process)
        dic_main['Process'][max(dic_main['SPC Item'].keys()) +1] = process
        cpk_dic['Process'][max(cpk_dic['SPC Item'].keys()) +1] = process
        process_lst .append(process)
        process_length_lst.append(max(cpk_dic['SPC Item'].keys()))
    print(dic_main['Process'])
    print(get_header_cell(mbo_pbo, build).coordinate)

    last_cell = get_column_letter(get_header_cell(mbo_pbo, build).column) + str(maxrow)
    m_mbo_pbo = mbo_pbo[m_data:last_cell]
    get_overall_dic(mbo_pbo, build, dic_main)
    #need to get the cells cells with the names in a dictionary with cpk values
    #use get_header_cell() to get the column but change the number
    #if "Liner" not in process:
        #get_overall_dic(mbo_pbo, build, dic_main)
    #else:
        #get_overall_liner_attach_dic(mbo_pbo, build, dic_main)
    #print(get_overall_dic(mbo_pbo, header, dic_main))
print(glob_header)
print(index_header)
process_length_lst.append(max(cpk_dic['SPC Item'].keys()))

cpk_lst = []
for ind in glob_header:
    if '5' in ind:
        cpk_lst.append(ind)
print(cpk_lst)

#ADD ANOTHER ONE FOR CPK VALUES ONLY
df2 = pd.DataFrame(cpk_dic, columns = cpk_head)

print(cpk_dic)
print(dic_main)
df = pd.DataFrame(dic_main, columns = glob_header)
print(df)
unique=datetime.datetime.now().strftime('%m%d%H%M') #gets current date and time in MMDDHHmm format MM = month, DD = date, HH = hour, mm = min
df.set_index(['Process', 'SPC Item'], inplace=True) #sets the 'Process' and 'SPC items' columns as index

new_name = direct + r'\new' + str(unique) + r'.xlsx'
writer = pd.ExcelWriter(new_name, engine='xlsxwriter')



sheetname = 'Sheet1'
df.to_excel(writer, sheet_name = sheetname) #new file is printed at where the original file path is found

workbook = writer.book
format1 = workbook.add_format({'bg_color': '#C6EFCE'}) #green
format2 = workbook.add_format({'bg_color': '#FFFFFF'}) #white
format3 = workbook.add_format({'bg_color': '#FFEB9C'}) #yellow
format4 = workbook.add_format({'bg_color': '#FF0000'}) #red
print(format)




worksheet = writer.sheets[sheetname]

for col in cpk_lst:

    print(df.columns.get_loc(col))
    letter = get_column_letter(df.columns.get_loc(col)+3) #the letter of the column of the cpk column
    print(letter)

    len_df = str(len(df.index) + 1)
    print(len_df)

    rng = letter + '3:' + letter + len_df #3 because we want to start from row 3 and end at the last row which is len_df
    print(rng)
    worksheet.conditional_format(rng, {'type': 'blanks','format': format2})
    worksheet.conditional_format(rng, {'type': 'text', 'criteria': 'containing', 'value': '-', 'format': format2})
    worksheet.conditional_format(rng, {'type': 'cell', 'criteria': '<', 'value': 1, 'format': format4}) #cpk<1.00 (red)
    worksheet.conditional_format(rng, {'type': 'cell', 'criteria': '<', 'value': 1.33, 'format': format3}) #1.00<cpk<1.33 (yellow)
    worksheet.conditional_format(rng, {'type': 'cell', 'criteria': '>', 'value': 1.33, 'format': format1}) #cpk>1.33 (green)

################# ADDING A NEW SHEET ##############################
sheetname2 = 'Overall_Cpk_Plot'
worksheet2=workbook.add_worksheet(sheetname2)
writer.sheets[sheetname2] = worksheet2
df2.set_index(['Process', 'SPC Item'], inplace=True)
df2.to_excel(writer, sheet_name = sheetname2)

(max_row_df2, max_col_df2) = df2.shape
print(max_col_df2)
letter_chart = get_column_letter(max_col_df2 +6) #letter of the column a few columns after the last data
worksheet = writer.sheets[sheetname2]

#plot
col = 2 #start col

#print(process_length_lst)
#print(process_lst)
max_col = len(cpk_head)-1

#print(max_col)

n_processes = len(process_length_lst) #number of processes e.g. cap attach, liner attach etc.
print(n_processes)
for n in range(n_processes): #iterating through the indexes of the number of processes
    #globals()['string%s' % x] allows you to create variable names through iteration
    globals()['chart%s' %n] =  workbook.add_chart({'type': 'line'})
    min_row = 2 #this is for the first process onwards
    if n != 0:
        min_row = process_length_lst[n-1] +2 #for the second process onwards
    if min_row >= process_length_lst[n]+2: #ensure that there is enough data to create a chart
        continue
    for r in range(min_row, process_length_lst[n]+2):
        globals()['chart%s' %n].add_series({ #adding data to the chart
            #format of [row, col] aka the cell that we want to pull the name from
            'name':       [sheetname2, r, 1],
            #format of [start_row, start_col, end_row, end col]
            'categories': [sheetname2, 0, col, 0, max_col], #obtains labels aka x axis 
            'values':     [sheetname2, r, col, r, max_col], #obtains data, aka y axis
        })
    globals()['chart%s' %n].set_title({'name': process_lst[n]}) #sets title of the charts
    globals()['chart%s' %n].set_x_axis({'name': 'Phases'}) #sets name of the x axis
    globals()['chart%s' %n].set_y_axis({'name': 'CPK', 'major_gridlines': {'visible': False}}) #sets name of the y axis
    globals()['chart%s' %n].set_table({'show_keys': True}) 

    # Insert the chart into the worksheet.
    ins = letter_chart + str(process_length_lst[n])  #grid where you want to insert your chart
    print(ins)
    worksheet.insert_chart(ins, globals()['chart%s' %n]) #insert chart into the excel at the specific cells

writer.save()
print('Done')


