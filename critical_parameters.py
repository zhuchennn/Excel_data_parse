from matplotlib import pyplot as plt
import openpyxl, os
import pandas as pd
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime
import tkinter as tk
import seaborn as sns

def loop_directory(direct: str):
    dic = {}
    for (root, dirs, files) in os.walk(direct, topdown=True):
        lst_files = []
        for i in files:
            if i.endswith('.xlsx') and not i.startswith('~$'):
                lst_files.append(i)
        dic[root] = lst_files
    return dic

def get_root_file(root, filename):
    return os.path.join(str(root), filename) #<---- This joins the directory(str) and filename(str) and returns a str

#input: a dictionary the keyword being the filename
#output: the root of the file and the filename in a 2 index list
def get_file_path(dic: dict, keyword: str):  #need to get a specific file with a specific input
    for root, filenames in dic.items():
        for filename in filenames:
            if keyword in filename:
                file_path = Path(get_root_file(root, filename))
                f = open(file_path)
                return [root, filename]

def get_key(dic: dict, val: str):
    for key, value in dic.items():
         if val == value:
             return key

window1 = tk.Tk()
canvas1 = tk.Canvas(window1, width=400, height=300, relief='raised')
canvas1.pack()

title = tk.Label(window1, text='Data Extraction Inputs')
title.config(font=('helvetica', 14))
canvas1.create_window(200, 25, window=title)

title1 = tk.Label(window1, text = 'File Path Input')
title1.config(font=('helvetica', 10))
canvas1.create_window(200, 80, window=title1)
direct = tk.Entry(window1)
canvas1.create_window(200, 115, window=direct)

title2 = tk.Label(window1, text= 'File Name Input')
title2.config(font=('helvetica', 10))
canvas1.create_window(200, 155, window=title2)
filename = tk.Entry(window1)
canvas1.create_window(200, 190, window=filename)

def close_window():
    global direct
    direct = direct.get()
    global filename
    filename = filename.get()
    window1.destroy()

button1 = tk.Button(text= 'confirm', command = close_window)
canvas1.create_window(200, 230, window=button1)

window1.mainloop()

######################################################################################################
direct = r'C:\Users\zzh2\OneDrive - Osram GmbH\Desktop\Zhu_Chen\Python_script\06211132_TestParameterPlot'
filename = 'meow.xlsx'
######################################################################################################
os.chdir(direct)

dic_files = loop_directory(direct)
root = get_file_path(dic_files, filename)[0]
file = get_file_path(dic_files, filename)[1] #actually redundant
data = pd.ExcelFile(file) #helps make the data from excel file readable by python
ps = openpyxl.load_workbook(file) #similar to above, makes excel file readable to python

sheet = ps[data.sheet_names[0]]
maxrow = sheet.max_row
maxcol = sheet.max_column
x = ps.active

#list of the name of files we want to run
lst_filenames = []
for row in x.iter_rows(max_col=1, max_row=maxrow):
    lst_filenames.append(row[0].value)

unique=datetime.datetime.now().strftime('%m%d%H%M')
new_name = direct + r'\Critical_Parameters_' + str(unique) + r'.xlsx'
writer = pd.ExcelWriter(new_name, engine='xlsxwriter')
sheetname = 'Sheet1'

workbook = writer.book
worksheet=workbook.add_worksheet(sheetname)

count = len(lst_filenames) * 12

x = 5
x2 = 0
for count in range(0, count, 12):
    imagename = lst_filenames[int(count/12)]
    worksheet.write(count,x2,imagename + '-Overall')
    worksheet.insert_image(count, x , imagename + '-Overall.png', {'x_scale': 0.5, 'y_scale': 0.5})

writer.save()
print('\nDone')

